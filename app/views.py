from django.shortcuts import render
from django.shortcuts import render
from django.utils.dateparse import parse_date
from datetime import datetime, timedelta
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models.functions import Coalesce
from django.db.models import Max, Min, F
from django.contrib.auth.models import Group, Permission
from rest_framework.permissions import IsAuthenticated, DjangoModelPermissions
from django.db.models import Prefetch 
from django.db.models import F, ExpressionWrapper, FloatField, OuterRef, Subquery
from rest_framework.decorators import action
from rest_framework.response import Response
from django.contrib.auth.models import User
from rest_framework import status
from rest_framework.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.utils.timezone import now 
from rest_framework.exceptions import ValidationError
from django.utils.dateparse import parse_datetime
from django.db import transaction
from rest_framework import viewsets
from django.utils.timezone import make_aware, is_naive
from django.db.models import Q
from django.db.models import Count, Sum
from django.utils.timezone import is_aware, make_aware
from rest_framework.filters import OrderingFilter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from .models import *
from .serializers import *


class EmailSettingsViewSet(viewsets.ModelViewSet):
    queryset = EmailSettings.objects.all()
    serializer_class = EmailSettingsSerializer

    

    @action(detail=False, methods=['post'], url_path='test_email')
    def test_email(self, request):

        email_settings = EmailSettings.objects.last()  
        if not email_settings:
            return Response({"error": "Email settings not configured."}, status=status.HTTP_400_BAD_REQUEST)
    
        subject = "Test Email"
        message = "perfect you are now ready to start .........."
        recipient_email = email_settings.email_host_user 

        try:
            send_mail(
                subject,
                message,
                email_settings.default_from_email, 
                [recipient_email]  
            )
            return Response({"status": "Test email sent successfully."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Failed to send test email: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class EmailSettingsAuditLogViewset(viewsets.ModelViewSet):
    queryset = EmailSettingsAuditLog.objects.all()
    serializer_class = EmailSettingsAuditLogSerializer

    def get_queryset(self):
        queryset = EmailSettingsAuditLog.objects.all()
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')

        if from_date:
            queryset = queryset.filter(change_time__date__gte=from_date) 
        if to_date:
            queryset = queryset.filter(change_time__date__lte=to_date)    

        return queryset




class EmailMessageViewSet(viewsets.ModelViewSet):
    queryset = EmailMessage.objects.all()
    serializer_class = EmailMessageSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        serializer.send_email()

        return Response(serializer.data, status=status.HTTP_201_CREATED)

class EmailGroupViewSet(viewsets.ModelViewSet):

    queryset = EmailGroup.objects.all()
    serializer_class = EmailGroupSerializer

    def update(self, request, *args, **kwargs):
        instance = self.get_object()  
        old_instance = EmailGroup.objects.get(pk=instance.pk) 
        serializer = self.get_serializer(instance, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save() 
            
            fields_to_check = ['name', 'arabic_name', 'emails']
            for field in fields_to_check:
                old_value = getattr(old_instance, field)
                new_value = getattr(instance, field)
                
                if old_value != new_value:
                    EmailGroupAuditLog.objects.create(
                        user=request.user, 
                        object_id=instance.pk,
                        field_name=field,
                        old_value=old_value,
                        new_value=new_value,
                        name=instance.name,
                        arabic_name=instance.arabic_name
                    )

            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class EmailGroupAuditLogViewSet(viewsets.ModelViewSet):
    serializer_class = EmailGroupAuditLogSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]

    def get_queryset(self):
        queryset = EmailGroupAuditLog.objects.all()
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')

        if from_date:
            queryset = queryset.filter(change_time__date__gte=from_date)
        if to_date:
            queryset = queryset.filter(change_time__date__lte=to_date)  

        return queryset
    

import logging

logger = logging.getLogger(__name__)

class EmailGroupExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        """Export EmailGroup data to Excel with emails in separate rows."""
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Email Groups"

            # Define headers
            headers = ["Group Name", "Arabic Name", "Email", "Notify"]
            ws.append(headers)

            # Add comment to 'Email' header from help_text
            emails_field = EmailGroup._meta.get_field('emails')
            ws['C1'].comment = Comment(emails_field.help_text, "API")

            # Fetch all EmailGroup instances
            email_groups = EmailGroup.objects.all()

            # Populate the sheet with one row per email
            for group in email_groups:
                email_list = [email.strip() for email in group.emails.split(',') if email.strip()]
                if not email_list:  # Handle empty emails field
                    ws.append([group.name, group.arabic_name or '', '', group.notify])
                else:
                    for email in email_list:
                        ws.append([group.name, group.arabic_name or '', email, group.notify])

            # Prepare the response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="email_groups.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"Error exporting Excel: {str(e)}")
            return Response(
                {"error": f"Failed to export Excel: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request, *args, **kwargs):
        """Import EmailGroup data from Excel with emails in separate rows."""
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            return Response(
                {"error": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not excel_file.name.endswith('.xlsx'):
            return Response(
                {"error": "Please upload an Excel file (.xlsx)."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Load the Excel file
            wb = load_workbook(excel_file)
            ws = wb.active

            # Expected headers
            expected_headers = ["Group Name", "Arabic Name", "Email", "Notify"]
            headers = [cell.value for cell in ws[1]]
            if headers != expected_headers:
                return Response(
                    {"error": "Invalid Excel format. Expected headers: Group Name, Arabic Name, Email, Notify"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Group emails by group name
            group_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                group_name, arabic_name, email, notify = row

                # Skip if group_name is missing
                if not group_name:
                    continue

                # Initialize group data if not already present
                if group_name not in group_data:
                    group_data[group_name] = {
                        'arabic_name': arabic_name if arabic_name else None,
                        'emails': [],
                        'notify': bool(notify) if notify is not None else False
                    }

                # Add email if present
                if email and email.strip():
                    group_data[group_name]['emails'].append(email.strip())

            # Process each group
            imported_count = 0
            for group_name, data in group_data.items():
                emails_str = ','.join(data['emails']) if data['emails'] else ''
                EmailGroup.objects.update_or_create(
                    name=group_name,
                    defaults={
                        'arabic_name': data['arabic_name'],
                        'emails': emails_str,
                        'notify': data['notify']
                    }
                )
                imported_count += 1

            return Response(
                {"message": f"Successfully imported {imported_count} email groups."},
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            logger.error(f"Error importing Excel: {str(e)}")
            return Response(
                {"error": f"Failed to import Excel: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )